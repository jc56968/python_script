# coding=utf-8
import openpyxl
import numpy as np

def insert(QQ,ss):

    for i in range(0,101):
        if(QQ[i]==0 ):
            j=i
            k=i
            while(j<100 and QQ[j]==0):
                j=j+1
            while (k>0 and QQ[k] == 0):
                k = k- 1
            if  QQ[k]!=0 and QQ[j]!=0:
                QQ[i]=QQ[k]+(QQ[j]-QQ[k])/(j-k)*(i-k)

        sheet4.cell(row=i+1, column=ss).value = QQ[i]
        sheet4.cell(row=i+1, column=ss + 1).value = i


wb = openpyxl.load_workbook('C:\\Users\\Administrator\\Documents\\WeChat Files\\aiyinsifeilun\Files\\45度不同放电倍率数据.xlsx')

# 从工作薄中获取一个表单(sheet)对象
sheets = wb.sheetnames
print(sheets, type(sheets))

# 创建一个表单
mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

# 获取指定的表单
sheet2 = wb.get_sheet_by_name('Sheet2')
ws=wb.active
print(ws)
s="BEHK"

for title in s:
        A=ws[title] # 获取第一行第二列的单元格
        for data in A:
            oldstr=data.value
            if oldstr!=None and oldstr!='SOC' :
                newstr=np.round(oldstr*100)/100
                print (newstr)
                data.value=newstr

wb.save("fi.xlsx")




wb = openpyxl.load_workbook("fi.xlsx")
ws=wb.active
sheet3 = wb.get_sheet_by_name('Sheet3')
sheet4 = wb.get_sheet_by_name('Sheet4')
print(ws)
s="BEHK"

start=18
end=10000
i=0
j=-2
big=0
small=0
count=0
if __name__=="__main__":
    for title in s:
            A=ws[title] # 获取第一行第二列的单元格
            B=ws[(chr(ord(title)-1))]
            j=j+3
            count=0
            end=10000
            i=0
            QQ=np.zeros(101)
            for data in A:
                oldstr=data.value
                count=count+1
                if oldstr != None and oldstr != 'SOC' and  oldstr>=0:
                    if end==10000:
                        end =count-1
                        start=oldstr
                        big=end
                        small=end
                    else:
                        print(oldstr,"PK")
                        if( start!=oldstr):
                            i=i+1
                            print(oldstr,count)
                            true_value=(B[big].value+B[small].value)/2
                            if(count==208):
                                count=count+0
                            sheet3.cell(row=i, column=j).value=true_value
                            sheet3.cell(row=i, column=j+1).value = A[end].value
                            QQ[int(A[end].value*100)]=true_value
                            print(QQ[0])
                            end=count-1
                            start=oldstr
                            big=end
                            small=end
                        elif B[count-1].value < B[small].value:
                            small=count-1

                        elif B[count-1].value >B[big].value:
                            big=count-1
            count = count + 1
            i=i+1
            true_value = (B[big].value + B[small].value) / 2
            sheet3.cell(row=i, column=j).value = true_value
            sheet3.cell(row=i, column=j + 1).value = A[end].value
            QQ[int(A[end].value * 100)] = true_value



            insert(QQ,j)

    wb.save("fi.xlsx")
    print(QQ[0])













########################################


